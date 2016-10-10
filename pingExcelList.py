from executePing import execPing
from getExcelColumnHeaders import getColumnHeaders
import openpyxl

def pingColumn(wb, ws, columnHeader, files, reset):

    if reset == True: permissions = "w"
    else: permissions = "a"

    successes = open(files["successFile"], permissions)
    timeouts = open(files["timeoutFile"], permissions)
    failures = open(files["failuresFile"], permissions)

    headers = getColumnHeaders(ws) #header dictionary

    for i in range(2, ws.max_row + 1):
        cell = (ws.cell(row=i, column = headers[columnHeader]))

        sep = "=========================================================" + "\n"
        if cell.value != columnHeader and cell.value != None:
            ping = execPing(hostName = cell.value, attempts = "2", waitTime = "3")
            if ping[0] == "Success: ":
                successes.write(ping[1] + "\n")
                successes.write(sep)
                print(sep)
                print(ping[1] + "\n")
            elif ping[0] == "Request timeout: ":
                timeouts.write(ping[1] + "\n")
                timeouts.write(sep)
                print(sep)
                print(ping[1] + "\n")
            elif ping[0] == "Failed: ":
                failures.write("*" + cell.value + "*" + ": " + ping[1] + "\n")
                failures.write(sep)
                print(sep)
                print("*" + cell.value + "*" + ": " + ping[1] + "\n")



if __name__ == "__main__":
    inFileName = "exampleWorkBook.xlsx"
    wb = openpyxl.load_workbook(inFileName, read_only=True) #load workbook
    ws = wb.get_sheet_by_name("Sheet1") #select worksheet

    cp = "_pings_from_"
    files = {"successFile" : "successful" + cp + inFileName.replace(".xlsx", ".txt"),
             "timeoutFile" : "timeout" + cp + inFileName.replace(".xlsx", ".txt"),
             "failuresFile" : "failure" + cp + inFileName.replace(".xlsx", ".txt")}

    pingColumn(wb = wb, ws = ws, columnHeader = "HostName", files = files, reset = True)
